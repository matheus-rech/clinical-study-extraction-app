"""
Export Service for Clinical Study Extraction App
Provides table exports (CSV, Excel, JSON) and annotated PDF generation
"""

import io
import json
import csv
from typing import List, Dict, Any, Optional
import logging
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fastapi import APIRouter, UploadFile, File, HTTPException, Response
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)
router = APIRouter()

class TableExporter:
    @staticmethod
    def to_csv(tables: List[Dict[str, Any]]) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        
        for idx, table in enumerate(tables):
            writer.writerow([f"=== Table {idx + 1} ==="])
            writer.writerow([f"Page: {table.get('page', 'N/A')}"])
            writer.writerow([])
            
            markdown = table.get('markdown', '')
            if markdown:
                for line in markdown.split('\n'):
                    if line.strip() and not line.strip().startswith('|---'):
                        cells = [c.strip() for c in line.strip('|').split('|')]
                        writer.writerow(cells)
            writer.writerow([])
        
        return output.getvalue()
    
    @staticmethod
    def to_excel(tables: List[Dict[str, Any]]) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)
        
        for idx, table in enumerate(tables):
            ws = wb.create_sheet(title=f"Table_{idx+1}")
            ws['A1'] = f"Table {idx + 1} - Page {table.get('page', 'N/A')}"
            ws['A1'].font = Font(bold=True, size=14)
            
            markdown = table.get('markdown', '')
            if markdown:
                row = 3
                for line in markdown.split('\n'):
                    if line.strip() and not line.strip().startswith('|---'):
                        cells = [c.strip() for c in line.strip('|').split('|')]
                        for col, cell in enumerate(cells, 1):
                            ws.cell(row=row, column=col, value=cell)
                        row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

class PDFAnnotator:
    @staticmethod
    def annotate_pdf(pdf_path: str, tables: List[Dict], figures: List[Dict] = None) -> bytes:
        doc = fitz.open(pdf_path)
        
        # Annotate tables
        for table in tables:
            page_num = table.get('page', 1) - 1
            if 0 <= page_num < len(doc):
                page = doc[page_num]
                bbox = table.get('bbox')
                if bbox:
                    rect = fitz.Rect(bbox['x0'], bbox['y0'], bbox['x1'], bbox['y1'])
                    page.draw_rect(rect, color=(0, 0, 1), width=2)
                    page.insert_text(
                        (rect.x0, rect.y0 - 5),
                        f"Table {table.get('table_index', 0) + 1}",
                        fontsize=10,
                        color=(0, 0, 1)
                    )
        
        output = io.BytesIO()
        doc.save(output)
        doc.close()
        output.seek(0)
        return output.getvalue()

@router.post("/api/export/tables/csv")
async def export_csv(tables: List[Dict[str, Any]]):
    csv_content = TableExporter.to_csv(tables)
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=tables.csv"}
    )

@router.post("/api/export/tables/excel")
async def export_excel(tables: List[Dict[str, Any]]):
    excel_content = TableExporter.to_excel(tables)
    return Response(
        content=excel_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tables.xlsx"}
    )

@router.post("/api/export/pdf/annotated")
async def export_annotated(file: UploadFile = File(...), tables: str = "[]"):
    tables_data = json.loads(tables)
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp:
        content = await file.read()
        temp.write(content)
        temp_path = temp.name
    
    annotated = PDFAnnotator.annotate_pdf(temp_path, tables_data)
    Path(temp_path).unlink()
    
    return Response(
        content=annotated,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=annotated_{file.filename}"}
    )

@router.post("/api/export/tables/json")
async def export_json(tables: List[Dict[str, Any]], pretty: bool = True):
    """Export tables to JSON format"""
    export_data = {
        "extraction_metadata": {
            "total_tables": len(tables),
            "export_format": "json",
            "version": "1.0"
        },
        "tables": tables
    }
    
    json_content = json.dumps(export_data, indent=2 if pretty else None, ensure_ascii=False)
    return Response(
        content=json_content,
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=tables.json"}
    )

@router.post("/api/export/tables/html")
async def export_html(tables: List[Dict[str, Any]]):
    """Export tables to HTML format with styling"""
    html_parts = [
        '<!DOCTYPE html>',
        '<html lang="en">',
        '<head>',
        '    <meta charset="UTF-8">',
        '    <meta name="viewport" content="width=device-width, initial-scale=1.0">',
        '    <title>Extracted Tables</title>',
        '    <style>',
        '        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }',
        '        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }',
        '        h1 { color: #333; border-bottom: 3px solid #4472C4; padding-bottom: 10px; }',
        '        .table-container { margin: 30px 0; }',
        '        .table-header { background: #4472C4; color: white; padding: 15px; margin-bottom: 10px; border-radius: 5px; }',
        '        .table-info { font-size: 14px; color: #666; margin-bottom: 10px; }',
        '        table { width: 100%; border-collapse: collapse; margin-bottom: 30px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }',
        '        th { background: #4472C4; color: white; padding: 12px; text-align: left; font-weight: bold; }',
        '        td { padding: 10px; border: 1px solid #ddd; }',
        '        tr:nth-child(even) { background: #f9f9f9; }',
        '        tr:hover { background: #f0f0f0; }',
        '        .metadata { background: #e7f3ff; padding: 15px; border-left: 4px solid #4472C4; margin-bottom: 20px; }',
        '    </style>',
        '</head>',
        '<body>',
        '    <div class="container">',
        '        <h1>📊 Extracted Tables Report</h1>',
        f'        <div class="metadata">',
        f'            <strong>Total Tables:</strong> {len(tables)}<br>',
        f'            <strong>Export Format:</strong> HTML<br>',
        f'            <strong>Generated:</strong> {json.dumps(tables[0].get("page") if tables else "N/A")}',
        '        </div>',
    ]
    
    for idx, table in enumerate(tables, 1):
        html_parts.extend([
            f'        <div class="table-container">',
            f'            <div class="table-header">',
            f'                <h2 style="margin:0;">Table {idx}</h2>',
            f'            </div>',
            f'            <div class="table-info">',
            f'                📄 Page: {table.get("page", "N/A")} | ',
            f'                📏 Dimensions: {table.get("rows", "N/A")} rows × {table.get("cols", "N/A")} cols | ',
            f'                🔧 Method: {table.get("extraction_method", "N/A")}',
            f'            </div>',
        ])
        
        # Convert markdown table to HTML
        markdown = table.get('markdown', '')
        if markdown:
            html_parts.append('            <table>')
            lines = markdown.split('\n')
            is_header = True
            
            for line in lines:
                if line.strip() and not line.strip().startswith('|---'):
                    cells = [c.strip() for c in line.strip('|').split('|')]
                    tag = 'th' if is_header else 'td'
                    html_parts.append('                <tr>')
                    for cell in cells:
                        html_parts.append(f'                    <{tag}>{cell}</{tag}>')
                    html_parts.append('                </tr>')
                    is_header = False
            
            html_parts.append('            </table>')
        
        html_parts.append('        </div>')
    
    html_parts.extend([
        '    </div>',
        '</body>',
        '</html>'
    ])
    
    html_content = '\n'.join(html_parts)
    return Response(
        content=html_content,
        media_type="text/html",
        headers={"Content-Disposition": "attachment; filename=tables.html"}
    )
